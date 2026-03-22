"""
Export unit test results to Excel.
Scans all *Test.java in src/test/java/.../service/ and merges with
surefire XML results to produce a detailed, formatted Excel report.
"""

import re
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── config ───────────────────────────────────────────────────────────────────
SERVICE_TEST_DIR = Path("src/test/java/com/swp391_be/SWP391_be/service")
SUREFIRE_DIR     = Path("target/surefire-reports")
OUTPUT_FILE      = "UnitTestReport_ServiceLayer.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────
def snake_to_words(name: str) -> str:
    """createBouquet_throwsWhenNameIsNull  →  'Throws when name is null'"""
    parts = name.split("_", 1)
    scenario = parts[1] if len(parts) > 1 else parts[0]
    # split camelCase
    words = re.sub(r"([A-Z])", r" \1", scenario).strip()
    return words[0].upper() + words[1:].lower() if words else name

def method_under_test(test_name: str) -> str:
    """createBouquet_throwsWhenNameIsNull  →  'createBouquet'"""
    return test_name.split("_")[0]

def extract_block(source: str, start_pos: int) -> str:
    """Extract the content of the { } block starting at start_pos."""
    depth = 0
    i = start_pos
    while i < len(source):
        if source[i] == '{':
            depth += 1
        elif source[i] == '}':
            depth -= 1
            if depth == 0:
                return source[start_pos: i + 1]
        i += 1
    return ""

def parse_precondition(body: str) -> str:
    """
    Extract human-readable preconditions from the test body.
    Looks for explicit value assignments and when() mocks.
    """
    lines = []

    # explicit field assignments like req.setName(null), req.setPrice(0), rawMaterial.setTotalQuantity(5)
    for m in re.finditer(r'(\w+)\.set(\w+)\(([^)]+)\)', body):
        obj, field, val = m.group(1), m.group(2), m.group(3).strip()
        # skip save/answer lambdas and mock setups
        if 'inv ->' in val or 'mock(' in val:
            continue
        lines.append(f"{obj}.{field[0].lower()+field[1:]} = {val}")

    # when(...).thenReturn(true/false/empty/...)
    for m in re.finditer(r'when\(([^)]+)\)\.thenReturn\(([^)]+)\)', body):
        call, ret = m.group(1).strip(), m.group(2).strip()
        if 'inv ->' not in ret:
            lines.append(f"mock: {call} → {ret}")

    # thenThrow
    for m in re.finditer(r'when\(([^)]+)\)\.thenThrow\(new (\w+)\(([^)]*)\)\)', body):
        call, exc, msg = m.group(1).strip(), m.group(2), m.group(3).strip('"')
        lines.append(f"mock: {call} throws {exc}({msg})")

    # deduplicate while preserving order
    seen, result = set(), []
    for l in lines:
        if l not in seen:
            seen.add(l)
            result.append(l)

    return "\n".join(result) if result else "Default / empty setup"

def parse_expected(body: str) -> str:
    """Extract expected result from assertions."""
    lines = []

    # assertThatThrownBy → exception + message
    thrown = re.search(
        r'assertThatThrownBy\(.+?\)\s*\.isInstanceOf\((\w+)\.class\)\s*\.hasMessageContaining\("([^"]+)"\)',
        body, re.DOTALL
    )
    if thrown:
        lines.append(f"Throws {thrown.group(1)}: \"{thrown.group(2)}\"")

    # assertThatCode → does not throw
    if re.search(r'assertThatCode\(.+?\)\.doesNotThrowAnyException\(\)', body, re.DOTALL):
        lines.append("No exception thrown")

    # assertThat(x).isEqualTo(y)
    for m in re.finditer(r'assertThat\(([^)]+)\)\.isEqualTo\(([^)]+)\)', body):
        lines.append(f"{m.group(1).strip()} == {m.group(2).strip()}")

    # assertThat(x).hasSize(n)
    for m in re.finditer(r'assertThat\(([^)]+)\)\.hasSize\((\d+)\)', body):
        lines.append(f"{m.group(1).strip()} has size {m.group(2)}")

    # assertThat(x).isEqualTo(y) chained  e.g. .isEqualTo(existingBouquet)
    for m in re.finditer(r'\.isEqualTo\(([^)]+)\)', body):
        val = m.group(1).strip()
        lines.append(f"result == {val}")

    # verify(repo).someMethod(...)
    for m in re.finditer(r'verify\((\w+)\)\.(\w+)\(', body):
        lines.append(f"verify {m.group(1)}.{m.group(2)}() called")

    # verify never
    for m in re.finditer(r'verify\((\w+),\s*never\(\)\)\.(\w+)\(', body):
        lines.append(f"verify {m.group(1)}.{m.group(2)}() never called")

    # deduplicate
    seen, result = set(), []
    for l in lines:
        if l not in seen:
            seen.add(l)
            result.append(l)

    return "\n".join(result) if result else "Assertion in body"

def parse_java_tests(java_file: Path) -> dict[str, dict]:
    """
    Returns {test_method_name: {method_under_test, description, precondition, expected}}
    """
    source = java_file.read_text(encoding="utf-8")
    tests  = {}

    # find all @Test annotated methods
    for m in re.finditer(r'@Test\s+(?:void\s+)?(\w+)\s*\([^)]*\)\s*(?:throws\s+\w+\s*)?\{', source):
        test_name = m.group(1)
        block     = extract_block(source, source.index('{', m.start()))
        tests[test_name] = {
            "method_under_test": method_under_test(test_name),
            "description":       snake_to_words(test_name),
            "precondition":      parse_precondition(block),
            "expected":          parse_expected(block),
        }
    return tests

# ── collect all service test classes ────────────────────────────────────────
all_tests = []   # list of dicts

java_files = sorted(SERVICE_TEST_DIR.glob("*Test.java"))
if not java_files:
    print(f"No *Test.java files found in {SERVICE_TEST_DIR}")
    raise SystemExit(1)

for java_file in java_files:
    class_name   = java_file.stem          # e.g. BouquetServiceTest
    xml_pattern  = f"TEST-*.{class_name}.xml"
    xml_files    = list(SUREFIRE_DIR.glob(xml_pattern))

    if not xml_files:
        # try broader pattern
        xml_files = [f for f in SUREFIRE_DIR.glob("TEST-*.xml") if class_name in f.name]

    parsed = parse_java_tests(java_file)

    if xml_files:
        tree = ET.parse(xml_files[0])
        root = tree.getroot()
        suite_elapsed = float(root.attrib.get("time", 0))

        for tc in root.findall("testcase"):
            name     = tc.attrib.get("name", "")
            time_sec = float(tc.attrib.get("time", 0))

            failure = tc.find("failure")
            error   = tc.find("error")
            skip    = tc.find("skipped")

            if failure is not None:
                status         = "FAILED"
                actual_result  = failure.attrib.get("message", failure.text or "See stack trace")
            elif error is not None:
                status         = "ERROR"
                actual_result  = error.attrib.get("message", error.text or "See stack trace")
            elif skip is not None:
                status         = "SKIPPED"
                actual_result  = skip.attrib.get("message", "Skipped")
            else:
                status         = "PASSED"
                actual_result  = "As expected"

            info = parsed.get(name, {})
            all_tests.append({
                "class":          class_name,
                "method_tested":  info.get("method_under_test", name.split("_")[0]),
                "description":    info.get("description", snake_to_words(name)),
                "precondition":   info.get("precondition", "-"),
                "expected":       info.get("expected", "-"),
                "actual":         actual_result,
                "status":         status,
                "time":           f"{time_sec:.3f}s",
            })
    else:
        # no XML yet — list from source only
        print(f"  WARNING: No surefire XML found for {class_name} – listing from source only")
        for test_name, info in parsed.items():
            all_tests.append({
                "class":          class_name,
                "method_tested":  info["method_under_test"],
                "description":    info["description"],
                "precondition":   info["precondition"],
                "expected":       info["expected"],
                "actual":         "Not run yet",
                "status":         "UNKNOWN",
                "time":           "-",
            })

# ── summary counts ───────────────────────────────────────────────────────────
total   = len(all_tests)
passed  = sum(1 for t in all_tests if t["status"] == "PASSED")
failed  = sum(1 for t in all_tests if t["status"] == "FAILED")
errors  = sum(1 for t in all_tests if t["status"] == "ERROR")
skipped = sum(1 for t in all_tests if t["status"] == "SKIPPED")

# ── Excel styles ─────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
LIGHT_BLUE = "D9E1F2"
WHITE      = "FFFFFF"
GREEN_BG   = "E2EFDA";  GREEN_FG   = "375623"
RED_BG     = "FFE0E0";  RED_FG     = "9C0006"
YELLOW_BG  = "FFF2CC";  YELLOW_FG  = "7D6608"
GRAY_BG    = "F2F2F2"

thin   = Side(style="thin",   color="AAAAAA")
medium = Side(style="medium", color="1F3864")
outer_border = Border(left=medium, right=medium, top=medium, bottom=medium)
inner_border = Border(left=thin,   right=thin,   top=thin,   bottom=thin)

def fill(color):
    return PatternFill("solid", fgColor=color)

def hdr_cell(ws, row, col, value, width=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=True, size=11, color=WHITE)
    c.fill      = fill(DARK_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = inner_border
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

# ── build workbook ───────────────────────────────────────────────────────────
wb = Workbook()

# ════════════════════════════════ OVERVIEW SHEET ════════════════════════════
ws_ov = wb.active
ws_ov.title = "Overview"

# title
ws_ov.merge_cells("A1:C1")
c = ws_ov["A1"]
c.value     = "Unit Test Report – Service Layer"
c.font      = Font(name="Calibri", bold=True, size=16, color=WHITE)
c.fill      = fill(DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_ov.row_dimensions[1].height = 32

# sub-title / date
import datetime
ws_ov.merge_cells("A2:C2")
c = ws_ov["A2"]
c.value     = f"Generated: {datetime.date.today().strftime('%Y-%m-%d')}"
c.font      = Font(name="Calibri", size=10, italic=True, color="555555")
c.fill      = fill(LIGHT_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_ov.row_dimensions[2].height = 16

# stat headers
for col, label in enumerate(["Metric", "Count", "Rate"], start=1):
    hdr_cell(ws_ov, 4, col, label)
ws_ov.row_dimensions[4].height = 20

stat_rows = [
    ("Total Tests",  total,   None),
    ("Passed",       passed,  f"{passed/total*100:.1f}%" if total else "-"),
    ("Failed",       failed,  f"{failed/total*100:.1f}%" if total else "-"),
    ("Errors",       errors,  f"{errors/total*100:.1f}%" if total else "-"),
    ("Skipped",      skipped, f"{skipped/total*100:.1f}%" if total else "-"),
]
for i, (label, count, rate) in enumerate(stat_rows, start=5):
    bg = GRAY_BG if i % 2 == 0 else WHITE
    for col, val in enumerate([label, count, rate or "-"], start=1):
        c = ws_ov.cell(row=i, column=col, value=val)
        c.font      = Font(name="Calibri", size=11,
                           bold=(label == "Total Tests"),
                           color=GREEN_FG if label == "Passed" and count > 0
                                 else RED_FG if label in ("Failed", "Errors") and count > 0
                                 else "000000")
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                vertical="center")
        c.border    = inner_border
    ws_ov.row_dimensions[i].height = 18

# per-class summary
row = len(stat_rows) + 7
ws_ov.merge_cells(f"A{row}:C{row}")
c = ws_ov.cell(row=row, column=1, value="Per-Class Summary")
c.font      = Font(name="Calibri", bold=True, size=12, color=WHITE)
c.fill      = fill(DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws_ov.row_dimensions[row].height = 22
row += 1

for col, label in enumerate(["Test Class", "Total", "Passed"], start=1):
    hdr_cell(ws_ov, row, col, label)
ws_ov.row_dimensions[row].height = 18
row += 1

classes = {}
for t in all_tests:
    cls = t["class"]
    classes.setdefault(cls, {"total": 0, "passed": 0})
    classes[cls]["total"] += 1
    if t["status"] == "PASSED":
        classes[cls]["passed"] += 1

for i, (cls, stats) in enumerate(classes.items()):
    bg = GRAY_BG if i % 2 == 0 else WHITE
    for col, val in enumerate([cls, stats["total"], stats["passed"]], start=1):
        c = ws_ov.cell(row=row, column=col, value=val)
        c.font      = Font(name="Calibri", size=10)
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
        c.border    = inner_border
    ws_ov.row_dimensions[row].height = 16
    row += 1

for col, w in zip(range(1, 4), [30, 10, 10]):
    ws_ov.column_dimensions[get_column_letter(col)].width = w

# ═══════════════════════════ DETAIL SHEET ════════════════════════════════════
ws = wb.create_sheet("Test Details")

# title
ws.merge_cells("A1:H1")
c = ws["A1"]
c.value     = "Unit Test Report – Detailed Results"
c.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
c.fill      = fill(DARK_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# summary bar
ws.merge_cells("A2:H2")
c = ws["A2"]
c.value = (f"Total: {total}   Passed: {passed}   Failed: {failed}   "
           f"Errors: {errors}   Skipped: {skipped}")
c.font      = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
c.fill      = fill(LIGHT_BLUE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# column headers + widths
COLUMNS = [
    ("#",                  5),
    ("Test Class",        22),
    ("Method Under Test", 22),
    ("Test Case Description", 42),
    ("Input / Precondition",  40),
    ("Expected Result",       38),
    ("Actual Result",         22),
    ("Status",                10),
]
for col_idx, (label, width) in enumerate(COLUMNS, start=1):
    hdr_cell(ws, 3, col_idx, label, width)
ws.row_dimensions[3].height = 22

# data rows
STATUS_STYLE = {
    "PASSED":  (GREEN_BG,  GREEN_FG),
    "FAILED":  (RED_BG,    RED_FG),
    "ERROR":   (RED_BG,    RED_FG),
    "SKIPPED": (YELLOW_BG, YELLOW_FG),
    "UNKNOWN": (GRAY_BG,   "555555"),
}

for row_idx, t in enumerate(all_tests, start=1):
    excel_row = row_idx + 3
    bg, fg = STATUS_STYLE.get(t["status"], (WHITE, "000000"))

    values = [
        row_idx,
        t["class"],
        t["method_tested"],
        t["description"],
        t["precondition"],
        t["expected"],
        t["actual"],
        t["status"],
    ]
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=excel_row, column=col_idx, value=val)
        c.fill      = fill(bg)
        c.border    = inner_border
        c.alignment = Alignment(
            horizontal="center" if col_idx in (1, 8) else "left",
            vertical="top",
            wrap_text=True
        )
        if col_idx == 8:  # Status column
            c.font = Font(name="Calibri", size=10, bold=True, color=fg)
        else:
            c.font = Font(name="Calibri", size=10)

    ws.row_dimensions[excel_row].height = 60  # tall enough for wrapped text

ws.freeze_panes = "A4"

# ── save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_FILE)
print(f"Done! {total} tests exported -> {OUTPUT_FILE}")
print(f"  Passed: {passed}  Failed: {failed}  Errors: {errors}  Skipped: {skipped}")
